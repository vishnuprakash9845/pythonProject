import openpyxl

#open the xl file(Workbook)
wb=openpyxl.open("/Users/vishnu/Downloads/Python-Selenium/Book1.xlsx")
sh=wb['Sheet2']
rc=sh.max_row
print("Row count",rc)
cc=sh.max_column
print("Column count",cc)
for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=sh.cell(i,j).value
        print(v," ",end='')
    print()

wb.close()