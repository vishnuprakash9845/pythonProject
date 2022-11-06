import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
#get user name and password from xl
wb=openpyxl.open("/Users/vishnu/Downloads/Python-Selenium/Book1.xlsx")
sheet3=wb['Sheet3']
un=sheet3.cell(2,1).value
pw=sheet3.cell(2,2).value

#open the browser and goto login page
driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()
driver.get("https://demo.actitime.com")
#enter un & pwd taken from xl & click login
driver.find_element(By.ID,"username").send_keys(un)
driver.find_element(By.NAME,"pwd").send_keys(pw)
driver.find_element(By.XPATH,"//div[.='Login ']").click()

#get the title of home page and write to xl file
wait=WebDriverWait(driver,5)
expected_title=sheet3.cell(2,3).value
print('Expected Title is:',expected_title)
try:
    wait.until(expected_conditions.title_is(expected_title))
    sheet3.cell(2, 5).value="PASS"
except:
    sheet3.cell(2, 5).value = "FAIL"

sheet3.cell(2, 4).value=driver.title
wb.save("/Users/vishnu/Downloads/Python-Selenium/Book1.xlsx")

wb.close()